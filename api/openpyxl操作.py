#用python操作excel
#第一步：导入load_workbook模块中的load_workbook函数
from openpyxl import load_workbook
#第二步：打开读取excel
wb=load_workbook(filename)#使用openpyxl读取xlsx文件，创建workbook,()里有多个参数，只有filename必填项
print(wb.sheetnames)#显示所有sheet名称，['sheet1名称', 'sheet2名称', ...]
#第三步：定位表单
ws=wb[sheetname]
#第四步：操作表单
#4.1定位单元格cell，根据行列读取数据
mm=ws.cell(row=2,column=3).value#读取某个sheet页第2行第3列单元格C2的值
#读取多个sheet的单元格
for i in wb.sheetnames:
    ws=wb[i]
    ws.cell(row=1,column=3).value#遍历所有的sheet页，读取每个sheet页C1的值
#4.2定位单元格（cell），根据行列值，更改原有的数据、写入新的测试数据，
mm="hello"#C2单元格写入'hello'
wb.save(filename)#wb.save('1.xlsx')保存文件
#4.3统计行列数
print(ws.max_row)#表行数
print(ws.max_column)#表列数
#类型
#excel 存储的数据，数字还是数字：int—>int、 float—>float 、其他类型—>str
#执行用例时需要将str()类型转化为原来的类型,eval(数据)
wb.close()#关闭

