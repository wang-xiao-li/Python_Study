import requests
import openpyxl
def post_requests(url,json):
    response=requests.post(url,json=json)
    return response.json()


def read_case(filename,sheetname):
    wk=openpyxl.load_workbook(filename)
    wb=wk[sheetname]
    cases=wb.max_row
    caselist=[]
    for i in range(2,cases+1):
        data1=dict(id=wb.cell(i,1).value,
                   url=wb.cell(i,2).value,
                   data=wb.cell(i,3).value,
                   expected=wb.cell(i,4).value)
        caselist.append(data1)
    return caselist


lists=read_case('case.xlsx','sheet1')
for list in lists:
    id=list.get('id')
    url=list.get('url')
    data=list.get('data')
    data=eval(data)
    expected=list.get('expected')
    global null
    null=''
    expected=eval(expected)#
    result=post_requests(url,data)
    print('实际结果：',result['data'])
    print('预期结果：',expected['data'])
    if result['data']==expected['data']:
        print('succeed')
        real_result='succeed'
    else:
        print('failed')
        real_result='failed'
    wk=openpyxl.load_workbook('case.xlsx')
    wb=wk['sheet1']
    wb.cell(id+1,5).value=str(result)
    wb.cell(id+1,6).value=real_result
    wk.save('case.xlsx')




