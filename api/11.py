import openpyxl
# book=openpyxl.Workbook()
# sheet=book.active
# sheet.title='爱情'
# sheet['A1']='id'
# sheet.cell(1,2).value='url'
# sheet.cell(1,3,value='data')
# book.save('case.xlsx')
bk=openpyxl.load_workbook('case.xlsx')
bs=bk['sheet1']
print(bs.max_row)